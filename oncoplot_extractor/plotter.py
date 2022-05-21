"Base file containing OncoPlotCreator"

from typing import DefaultDict
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from .errors import OncoPlotPlotterError


class OncoplotCreator:
    """Creates an oncoplot from a dataframe
    containing respective rows and columns with hexcolor data

    :param df: pandas dataframe containing hexcolor data [pandas.DataFrame]
    :param cell_size: size of each cell in pixels [int]
    :param workbook: openpyxl workbook object [openpyxl.Workbook]
    :param offset: number of rows to offset inserted data [int]

    :Example:
    .. code-block:: python

        opc = OncoplotCreator(df=my_df, cell_size=60, workbook=wb, offset=10)
        opc.gen_base_oncoplot()
        opc.save(filename="my_oncoplot.xlsx")

    """

    CELL_SIZE: int
    OUT_PATH: str = None
    DF: pd.DataFrame
    WORKBOOK: openpyxl.Workbook
    WORKSHEET: openpyxl.Workbook.active
    OFFSET: int

    def __init__(
        self,
        df: pd.DataFrame,
        cell_size: int = 60,
        workbook: openpyxl.Workbook = openpyxl.Workbook(),
        offset: int = 10,
    ) -> None:
        """Initializes a OncoplotCreator class, expects `df`, `cell_size`, `workbook`, `offset`
        which is row offset inserted at beginning of self.WS"""
        self.DF = df
        self.CELL_SIZE = cell_size
        self.WORKBOOK = workbook
        self.WORKSHEET = self.WORKBOOK.active
        self.OFFSET = offset
        self._load_df_to_ws()

    def _load_df_to_ws(self):
        """loads self.DF to self.WS"""
        for row in dataframe_to_rows(self.DF, index=True):
            self.WORKSHEET.append(row)

    def _insert_offset(self):
        for i in range(0, self.OFFSET):
            self.WORKSHEET.insert_rows(0)

    def gen_base_oncoplot(self):
        """Generates base oncoplot from self.DF
        expects self.DF to be loaded to self.WS
        """
        thin = openpyxl.styles.Side(border_style="thin", color="ffffff")
        self._insert_offset()
        _c = 1
        for row in self.WORKSHEET.iter_rows():
            self.WORKSHEET.row_dimensions[_c].height = self.CELL_SIZE
            for cell in row:
                if type(cell.value) == str:
                    cell.fill = openpyxl.styles.PatternFill(
                        patternType="solid", fgColor=cell.value.replace("#", "")
                    )
                    cell.value = ""
                    cell.border = openpyxl.styles.Border(
                        top=thin, bottom=thin, right=thin, left=thin
                    )
            _c += 1

    def gen_stack_plot(
        self,
        row_position: int = None,
        column_position: int = 2,
        filter_colors: list = [],
    ) -> None:
        """generates stacked barplot on top of base oncoplot
        :param row_position: row where the base of stackplot lies [int] (default: None)
        :param column_position: first column position [int] (default: 2)
        :param filter_colors: list of colors to filter out [list]"""
        col_pos = column_position
        plot_data = DefaultDict()
        thin = openpyxl.styles.Side(border_style="thin", color="ffffff")

        for col in self.DF.columns:
            counts = DefaultDict()
            for val in self.DF[col]:
                if not val in counts.keys():
                    counts[val] = 1
                else:
                    counts[val] += 1
            plot_data[col] = counts
            if row_position is None:
                row_pos = max(counts.items(), key=lambda k: k[1])[1] + 1
            else:
                row_pos = row_position
            c = 0
            for val in counts.keys():
                if val in filter_colors:
                    continue
                for i in range(1, counts[val] + 1):
                    try:
                        cell = self.WORKSHEET.cell(row=row_pos, column=col_pos)
                    except ValueError:
                        raise OncoPlotPlotterError(
                            f"Column or row position out of range col={col_pos}, row={row_pos}"
                        )
                    fgcolor = val.replace("#", "")
                    print(fgcolor)
                    cell.fill = openpyxl.styles.PatternFill(
                        patternType="solid",
                        fgColor=fgcolor,
                    )
                    cell.value = ""
                    cell.border = openpyxl.styles.Border(
                        top=thin, bottom=thin, right=thin, left=thin
                    )
                    row_pos -= 1
                c += 1
            col_pos += 1

    def save(self, filename: str) -> None:
        """saves generated workbook to file
        :param filename: filename to save workbook to [str]
        """
        self.WORKBOOK.save(filename=filename)
